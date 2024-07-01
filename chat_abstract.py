# -*- coding: utf-8 -*-
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
import re

# 指定文件路径
file_path = '/Users/songyujian/Downloads/聊天记录.xlsx'
chunk_size = 100  # 每次处理100行

pd.set_option('display.max_rows', None) # 设置显示最大行
pd.set_option('display.max_columns', None) # 设置显示最大列
pd.set_option('display.width', None) # 设置输出宽度
pd.set_option('display.max_colwidth', None) # 设置列最大宽度

wb = load_workbook(filename=file_path, read_only=True)
sheet = wb.active
total_rows = sheet.max_row - 1  # 减去标题行
wb.close()

# 文件中总共有多少行
# 使用`nrows`来读取Excel文件的第一行以获取列名
df_temp = pd.read_excel(file_path, nrows=1)

# 计算需要多少个chunk（向上取整）
num_chunks = -(-total_rows // chunk_size)

# 设置AI检查的API客户端
client = OpenAI(
    api_key="sk-l6r0OoY09uVet8ybar31YrnIpQaJ0DiDsn4z8enRVTEqUslc",  # 替换为你的Moonshot API密钥
    base_url="https://api.moonshot.cn/v1",
)

user = '姚'


# 返回AI判断
def get_summary(mydf):
    # 发送问题到Moonshot AI
    completion = client.chat.completions.create(
        model="moonshot-v1-8k",
        messages=[
            {"role": "system",
             "content": f"你是一个中文的群聊发言人信息总结的助手，你可以为一个发言人在微信的群聊记录，提取并总结每个时间段{user}参与讨论的话题内容。忽略其中的[图片]、[动画表情]类发言。"},
            {"role": "user", "content": f"请帮我将给出的群聊内容总结成一个分日期的发言人{user}在群聊的活跃报告，包含{user}参与的所有话题的总结。"
                                        f"每个话题的格式必须为：1、话题名(50字以内⃣）2、参与者(必须包含发送者{user}) 3、时间段(从几点到几点) 4、过程(50到200字左右）5、对{user}发言可能回复的提问 6、{user}聊天消息原文。"
                                        f"另外有以下要求：1.使用中文冒号2.无需大标题。群聊记录为{mydf}"}
        ]
        ,
        temperature=0.3,
    )
    # 获取并返回回答内容
    response = completion.choices[0].message.content
    return response


# 使用正则表达式匹配所有以数字编号开头的段落，假设每个话题都是以"1、"开始
def gen_textblock(summary_text):
    topic_blocks = re.split(r'\n(?=\d+、)', summary_text.strip())
    return topic_blocks


# 清理特殊字符函数保持不变
def clean_special_characters(topic_blocks):
    cleaned_blocks = []
    # 定义一个正则表达式模式来匹配所有非打印字符和Unicode空格
    pattern = re.compile(r'[\u2000-\u206F\u2E00-\u2E7F\\\'!"#$%&()*+,\-.\/:;<=>?@\[\]^_`{|}~]')

    for block in topic_blocks:
        # 使用正则表达式替换特殊字符为空格
        cleaned_block = pattern.sub(' ', block)
        # 将清理后的文本添加到新列表中
        cleaned_blocks.append(cleaned_block)

    return cleaned_blocks


def format_res(cleaned_topic_blocks, topics, participants, time_ranges, processes, questions, original_texts):
    # 遍历所有匹配到的话题块
    topic_info = {'1': '', '2': '', '3': '', '4': '', '5': '', '6': ''}
    for item in cleaned_topic_blocks:
        # 对于以 "1、" 开头的项，只提取冒号前面的内容作为话题名
        if item.startswith('1、'):
            # topics.append(item[2:])
            topic_info['1'] = item[2:]
        else:
            # 对于其他项，提取冒号后面的内容
            match = re.search(r'(?:\d、.*：)(.*)', item)
            if match:
                content = match.group(1).strip()  # 移除前缀和任何额外的空白字符
                if item.startswith('2、'):
                    # participants.append(content)
                    topic_info['2'] = content
                elif item.startswith('3、'):
                    # time_ranges.append(content)
                    topic_info['3'] = content
                elif item.startswith('4、'):
                    # processes.append(content)
                    topic_info['4'] = content
                elif item.startswith('5、'):
                    # questions.append(content)
                    topic_info['5'] = content
                elif item.startswith('6、'):
                    # original_texts.append(content)
                    topic_info['6'] = content
                # 将提取到的信息添加到对应的列表中
        topics.append(topic_info['1'])
        participants.append(topic_info['2'])
        time_ranges.append(topic_info['3'])
        processes.append(topic_info['4'])
        questions.append(topic_info['5'])
        original_texts.append(topic_info['6'])




all_topics, all_participants, all_time_ranges, all_processes, all_questions, all_original_texts = [], [], [], [], [], []

for chunk_index in range(3):
    # 计算当前chunk的起始位置
    chunk_start = chunk_index * chunk_size + 1  # 加1是因为跳过标题行

    # 使用`skiprows`跳过之前已经读取的行，但是要保留列名
    # 使用`nrows`来限制每次读取的行数
    df_chunk = pd.read_excel(
        file_path,
        skiprows=chunk_start,  # 跳过不需要的行
        nrows=chunk_size  # 指定读取行数
    )

    # 获取当前数据块的摘要文本
    summary_text = get_summary(df_chunk)
    topic_blocks = gen_textblock(summary_text)
    # 清理特殊字符
    cleaned_topic_blocks = clean_special_characters(topic_blocks)


    # 处理摘要文本，提取相关信息，并添加到各自的列表中
    topics, participants, time_ranges, processes, questions, original_texts = [], [], [], [], [], []
    format_res(cleaned_topic_blocks, topics, participants, time_ranges, processes, questions, original_texts)

    # 将当前块处理后的信息添加到总列表中
    all_topics.extend(topics)
    all_participants.extend(participants)
    all_time_ranges.extend(time_ranges)
    all_processes.extend(processes)
    all_questions.extend(questions)
    all_original_texts.extend(original_texts)

print(all_topics)
print(all_questions)
# 创建汇总DataFrame
df_summary = pd.DataFrame({
    '话题': all_topics,
    '参与者': all_participants,
    '时间段': all_time_ranges,
    '过程': all_processes,
    '可能回复的提问': all_questions,
    '聊天消息原文': all_original_texts
})

# 将DataFrame存储到新的Excel文件中
output_file_path = '/Users/songyujian/Downloads/whole_summary_output.xlsx'
df_summary.to_excel(output_file_path, index=False)
